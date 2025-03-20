# -*- coding: utf-8 -*-
"""
@Project : AICode
@File    : app_ui.py
@Author  : xuan
@Date    : 2025/3/7 17:16
"""
import asyncio
import streamlit as st
import os
import time
import json
import pandas as pd
import io
import datetime
from pydantic import BaseModel, Field
from typing import List, Dict, Any
import fitz  # PyMuPDF
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from autogen_agentchat.agents import AssistantAgent
from llms import model_client
from prompt_tasks import TESTCASE_WRITER_SYSTEM_MESSAGE

# 设置页面配置
st.set_page_config(
    page_title="上传需求文档&测试用例生成器",
    page_icon="✅",
    layout="wide"
)

# 确保data目录存在
os.makedirs("data", exist_ok=True)

# 页面标题
st.title("🧪 AI 上传需求文档&测试用例生成器")
st.markdown("上传需求文档或输入需求描述，AI 将为你生成相应的测试用例")


# 定义测试用例模型
class TestCase(BaseModel):
    case_id: str = Field(..., description="测试用例唯一标识符，格式为TC-XXX-NNN")
    priority: str = Field(..., description="优先级，P0(最高)、P1、P2、P3(最低)")
    title: str = Field(..., description="测试用例标题")
    precondition: str = Field(..., description="前置条件")
    steps: str = Field(..., description="测试步骤")
    expected_result: str = Field(..., description="预期结果")


class TestCaseCollection(BaseModel):
    test_cases: List[TestCase] = Field(..., description="测试用例集合")


# 初始化会话状态
if 'extracted_text' not in st.session_state:
    st.session_state.extracted_text = ""
if 'generated_testcases' not in st.session_state:
    st.session_state.generated_testcases = None
if 'processing_time' not in st.session_state:
    st.session_state.processing_time = {"extraction": 0, "generation": 0}
if 'uploaded_file_path' not in st.session_state:
    st.session_state.uploaded_file_path = None


# 创建测试用例生成器代理
@st.cache_resource
def get_testcase_writer():
    return AssistantAgent(
        name="testcase_writer",
        model_client=model_client,
        system_message=TESTCASE_WRITER_SYSTEM_MESSAGE,
        model_client_stream=True,
    )


# 从PDF提取文本
def extract_text_from_pdf(pdf_file) -> Dict[str, Any]:
    """从PDF文件中提取文本内容"""
    start_time = time.time()

    try:
        # 直接从内存加载PDF
        pdf_bytes = pdf_file.getvalue()
        doc = fitz.open(stream=pdf_bytes, filetype="pdf")

        # 提取文本
        text_content = ""
        metadata = {
            "title": doc.metadata.get("title", "未知"),
            "author": doc.metadata.get("author", "未知"),
            "subject": doc.metadata.get("subject", "未知"),
            "pages": len(doc),
        }

        for page_num in range(len(doc)):
            page = doc[page_num]
            text_content += page.get_text()

        elapsed_time = time.time() - start_time
        return {
            "text": text_content,
            "metadata": metadata,
            "elapsed_time": elapsed_time
        }
    except Exception as e:
        st.error(f"PDF处理错误: {str(e)}")
        return {
            "text": "",
            "metadata": {"error": str(e)},
            "elapsed_time": time.time() - start_time
        }


# 解析测试用例JSON
def parse_testcases_from_json(json_str: str) -> TestCaseCollection:
    """从JSON字符串解析测试用例列表"""
    try:
        # 清理可能的前缀和后缀文本
        json_text = json_str.strip()

        # 查找JSON开始和结束的位置
        start_idx = json_text.find('{')
        end_idx = json_text.rfind('}')

        if start_idx >= 0 and end_idx > start_idx:
            json_text = json_text[start_idx:end_idx + 1]

        # 尝试直接解析JSON
        data = json.loads(json_text)

        # 如果JSON是一个字典，并且有"test_cases"字段
        if isinstance(data, dict) and "test_cases" in data:
            return TestCaseCollection.model_validate(data)

        # 如果JSON是一个列表，假设它是测试用例列表
        if isinstance(data, list):
            return TestCaseCollection(test_cases=data)

        # 其他情况，尝试包装为测试用例列表
        return TestCaseCollection(test_cases=[data])

    except Exception as e:
        st.error(f"解析JSON测试用例时出错: {str(e)}")

        # 尝试使用更宽松的方式提取JSON
        try:
            # 使用正则表达式查找所有可能的JSON对象
            import re
            json_objects = re.findall(r'\{(?:[^{}]|(?:\{[^{}]*\}))*\}', json_str)

            if json_objects:
                # 尝试解析找到的第一个完整JSON对象
                for json_obj in json_objects:
                    try:
                        data = json.loads(json_obj)
                        if "case_id" in data:
                            return TestCaseCollection(test_cases=[data])
                        elif "test_cases" in data:
                            return TestCaseCollection.model_validate(data)
                    except:
                        continue

            # 尝试查找JSON数组
            json_arrays = re.findall(r'\[(?:[^\[\]]|(?:\[[^\[\]]*\]))*\]', json_str)
            if json_arrays:
                for json_array in json_arrays:
                    try:
                        data = json.loads(json_array)
                        if isinstance(data, list) and len(data) > 0:
                            return TestCaseCollection(test_cases=data)
                    except:
                        continue
        except:
            pass

        # 如果所有尝试都失败，手动解析
        try:
            # 查找测试用例模式
            test_cases = []
            case_blocks = re.findall(r'case_id["\']?\s*:\s*["\']?TC-[^"\'}\s]+["\']?', json_str)

            for i, start in enumerate(case_blocks):
                # 创建一个基本测试用例
                test_case = {
                    "case_id": f"TC-GEN-{i + 1:03d}",
                    "priority": "P1",
                    "title": f"自动生成测试用例 {i + 1}",
                    "precondition": "系统正常运行",
                    "steps": "1. 执行测试步骤",
                    "expected_result": "符合预期的结果"
                }
                test_cases.append(test_case)

            if test_cases:
                return TestCaseCollection(test_cases=test_cases)
        except:
            pass

        # 返回空列表作为默认值
        return TestCaseCollection(test_cases=[])


# 修改导出JSON的部分
def export_json(testcases: TestCaseCollection) -> str:
    """将测试用例导出为JSON字符串"""
    try:
        # 尝试使用 Pydantic v2 的方法
        if hasattr(testcases, 'model_dump_json'):
            try:
                # 尝试使用 ensure_ascii=False
                return testcases.model_dump_json(indent=2, ensure_ascii=False)
            except TypeError:
                # 如果不支持 ensure_ascii 参数
                return testcases.model_dump_json(indent=2)
        # 兼容 Pydantic v1
        elif hasattr(testcases, 'json'):
            return testcases.json(indent=2, ensure_ascii=False)
        else:
            # 手动创建JSON
            test_cases_dict = {"test_cases": []}
            for tc in testcases.test_cases:
                test_cases_dict["test_cases"].append({
                    "case_id": tc.case_id,
                    "priority": tc.priority,
                    "title": tc.title,
                    "precondition": tc.precondition,
                    "steps": tc.steps,
                    "expected_result": tc.expected_result
                })
            return json.dumps(test_cases_dict, indent=2, ensure_ascii=False)
    except Exception as e:
        st.error(f"导出JSON时出错: {str(e)}")
        # 最后的备选方案：使用最基本的方式
        try:
            result = []
            for tc in testcases.test_cases:
                result.append({
                    "case_id": tc.case_id,
                    "priority": tc.priority,
                    "title": tc.title,
                    "precondition": tc.precondition,
                    "steps": tc.steps,
                    "expected_result": tc.expected_result
                })
            return json.dumps({"test_cases": result}, indent=2)
        except:
            return "{\"test_cases\": []}"


# 生成测试用例
async def generate_testcases(requirement_text: str, test_level: str, test_priority: str, test_case_count: int) -> str:
    """生成测试用例"""
    testcase_writer = get_testcase_writer()

    # 构建提示信息
    prompt = f"""
    请根据以下需求描述生成测试用例:

    需求描述:
    {requirement_text}

    测试级别: {test_level}
    测试优先级: {test_priority}
    测试用例数量: 请生成 {test_case_count} 个测试用例

    请确保测试用例涵盖主流程、边界条件和异常情况。
    请以JSON格式输出，包含测试用例ID、优先级、标题、前置条件、测试步骤和预期结果。
    输出格式必须是有效的JSON，格式如下:
    {{
      "test_cases": [
        {{
          "case_id": "TC-XXX-001",
          "priority": "P1",
          "title": "测试用例标题",
          "precondition": "前置条件",
          "steps": "测试步骤",
          "expected_result": "预期结果"
        }},
        ...
      ]
    }}
    """

    # 使用正确的异步流式调用
    full_response = ""
    try:
        async for chunk in testcase_writer.run_stream(task=prompt):
            if hasattr(chunk, 'content'):
                full_response += chunk.content
            elif isinstance(chunk, str):
                full_response += chunk
            else:
                full_response += str(chunk)
        return full_response
    except Exception as e:
        st.error(f"生成失败: {str(e)}")
        return """
        {
          "test_cases": [
            {
              "case_id": "TC-ERR-001",
              "priority": "P1",
              "title": "生成失败 - 请重试",
              "precondition": "系统正常运行",
              "steps": "1. 尝试生成测试用例",
              "expected_result": "生成有效的测试用例"
            }
          ]
        }
        """


# 导出为Excel
def export_to_excel(testcases: TestCaseCollection) -> bytes:
    """将测试用例导出为Excel文件"""
    # 创建工作簿和工作表
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "测试用例"

    # 设置列宽
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 40
    ws.column_dimensions['F'].width = 40

    # 设置表头
    headers = ["用例ID", "优先级", "标题", "前置条件", "测试步骤", "预期结果"]
    header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    header_font = Font(bold=True)

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # 填充数据
    for row_num, testcase in enumerate(testcases.test_cases, 2):
        ws.cell(row=row_num, column=1, value=testcase.case_id)
        ws.cell(row=row_num, column=2, value=testcase.priority)
        ws.cell(row=row_num, column=3, value=testcase.title)
        ws.cell(row=row_num, column=4, value=testcase.precondition)
        ws.cell(row=row_num, column=5, value=testcase.steps)
        ws.cell(row=row_num, column=6, value=testcase.expected_result)

        # 设置单元格对齐方式
        for col_num in range(1, 7):
            ws.cell(row=row_num, column=col_num).alignment = Alignment(wrap_text=True, vertical='top')

    # 保存到内存
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return output.getvalue()


# 导出为Markdown
def export_to_markdown(testcases: TestCaseCollection) -> str:
    """将测试用例导出为Markdown格式"""
    md_content = "# 测试用例集\n\n"
    md_content += f"生成时间: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n"

    for i, testcase in enumerate(testcases.test_cases, 1):
        md_content += f"## {i}. {testcase.title} ({testcase.case_id})\n\n"
        md_content += f"- **优先级**: {testcase.priority}\n"
        md_content += f"- **前置条件**: {testcase.precondition}\n\n"
        md_content += "### 测试步骤\n\n"
        md_content += f"{testcase.steps}\n\n"
        md_content += "### 预期结果\n\n"
        md_content += f"{testcase.expected_result}\n\n"
        md_content += "---\n\n"

    return md_content


# 创建选项卡
source_tab1, source_tab2 = st.tabs(["📄 文档上传", "✏️ 手动输入"])

with source_tab1:
    uploaded_file = st.file_uploader("上传需求文档", type=["pdf"])

    if uploaded_file is not None:
        # 保存上传的文件到data目录
        file_path = os.path.join("data", uploaded_file.name)
        with open(file_path, "wb") as f:
            f.write(uploaded_file.getvalue())

        st.session_state.uploaded_file_path = file_path
        st.success(f"文件已上传并保存到: {file_path}")

        # 提取文本
        extraction_result = extract_text_from_pdf(uploaded_file)
        st.session_state.extracted_text = extraction_result["text"]
        st.session_state.processing_time["extraction"] = extraction_result["elapsed_time"]

        # 显示文档元数据
        with st.expander("文档元数据"):
            st.json(extraction_result["metadata"])

        # 显示提取的文本
        with st.expander("提取的文本内容"):
            st.text_area("文本内容", st.session_state.extracted_text, height=300)

with source_tab2:
    manual_input = st.text_area(
        "需求描述",
        height=300,
        placeholder="请详细描述你的功能需求，例如：\n开发一个用户注册功能，要求用户提供用户名、密码和电子邮件。用户名长度为3-20个字符，密码长度至少为8个字符且必须包含数字和字母，电子邮件必须是有效格式。",
        key="manual_requirements_input"
    )

    if manual_input:
        st.session_state.extracted_text = manual_input

# 高级选项（可折叠）
with st.expander("高级选项"):
    col1, col2, col3 = st.columns(3)

    with col1:
        test_level = st.selectbox(
            "测试级别",
            ["单元测试", "集成测试", "系统测试", "验收测试"],
            index=2
        )

    with col2:
        test_priority = st.selectbox(
            "测试优先级",
            ["高", "中", "低"],
            index=0
        )

    with col3:
        # 添加测试用例数量控制
        test_case_count = st.number_input(
            "生成测试用例数量",
            min_value=3,
            max_value=100,
            value=10,
            step=1
        )

# 处理按钮
generate_button = st.button("生成测试用例", use_container_width=True,
                            disabled=not st.session_state.extracted_text)

# 生成测试用例
if generate_button and st.session_state.extracted_text:
    with st.spinner("正在生成测试用例..."):
        start_time = time.time()

        # 创建进度条
        progress_bar = st.progress(0)
        progress_text = st.empty()

        # 更新进度
        progress_bar.progress(10)
        progress_text.text("正在分析需求...")

        # 异步生成测试用例
        try:
            # 执行异步操作
            loop = asyncio.new_event_loop()
            asyncio.set_event_loop(loop)
            testcase_json = loop.run_until_complete(
                generate_testcases(
                    st.session_state.extracted_text,
                    test_level,
                    test_priority,
                    test_case_count
                )
            )
            loop.close()

            # 更新进度
            progress_bar.progress(70)
            progress_text.text("正在解析测试用例...")

            # 解析生成的测试用例
            testcases = parse_testcases_from_json(testcase_json)
            st.session_state.generated_testcases = testcases

            # 记录处理时间
            elapsed_time = time.time() - start_time
            st.session_state.processing_time["generation"] = elapsed_time

            # 更新进度
            progress_bar.progress(100)
            progress_text.text(f"测试用例生成完成，耗时: {elapsed_time:.2f}秒")

            # 显示成功消息
            st.success(f"成功生成 {len(testcases.test_cases)} 个测试用例")

        except Exception as e:
            st.error(f"生成测试用例时出错: {str(e)}")

# 显示生成的测试用例
if st.session_state.generated_testcases:
    st.subheader("生成的测试用例")

    # 创建表格显示测试用例
    testcases_data = []
    for tc in st.session_state.generated_testcases.test_cases:
        testcases_data.append({
            "用例ID": tc.case_id,
            "优先级": tc.priority,
            "标题": tc.title,
            "前置条件": tc.precondition,
            "测试步骤": tc.steps,
            "预期结果": tc.expected_result
        })

    df = pd.DataFrame(testcases_data)
    st.dataframe(df, use_container_width=True)

    # 导出选项
    st.subheader("导出测试用例")
    col1, col2, col3 = st.columns(3)

    # 导出为Excel
    with col1:
        excel_data = export_to_excel(st.session_state.generated_testcases)
        st.download_button(
            label="下载Excel文件",
            data=excel_data,
            file_name=f"测试用例_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    # 导出为JSON
    with col2:
        json_data = export_json(st.session_state.generated_testcases)
        st.download_button(
            label="下载JSON文件",
            data=json_data,
            file_name=f"测试用例_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.json",
            mime="application/json"
        )

    # 导出为Markdown
    with col3:
        md_data = export_to_markdown(st.session_state.generated_testcases)
        st.download_button(
            label="下载Markdown文件",
            data=md_data,
            file_name=f"测试用例_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.md",
            mime="text/markdown"
        )

# 显示处理时间统计
if any(st.session_state.processing_time.values()):
    with st.expander("处理时间统计"):
        times = st.session_state.processing_time
        time_df = pd.DataFrame({
            "步骤": ["文本提取", "测试用例生成"],
            "耗时(秒)": [times["extraction"], times["generation"]]
        })
        st.dataframe(time_df)

# 添加使用说明
with st.sidebar.expander("使用说明", expanded=True):
    st.markdown("""
    ### 使用步骤
    1. 上传PDF文档或手动输入需求描述
    2. 设置高级选项（可选）
    3. 点击"生成测试用例"按钮
    4. 查看生成的测试用例
    5. 下载所需格式的测试用例文件

    ### 高级选项说明
    - **测试级别**：选择测试用例的级别（单元、集成、系统或验收）
    - **测试优先级**：设置整体测试优先级
    - **测试用例数量**：控制生成的测试用例数量
    """)

# 添加关于信息
with st.sidebar.expander("关于"):
    st.markdown("""
    ### AI 测试用例生成器

    本工具使用人工智能技术，根据需求描述自动生成测试用例。

    **特点**：
    - 支持PDF文档上传和文本输入
    - 自动提取需求信息
    - 生成结构化的测试用例
    - 支持多种格式导出（Excel、JSON、Markdown）

    **技术栈**：
    - Streamlit：前端界面
    - AutoGen：AI代理框架
    - PyMuPDF：PDF处理
    - OpenPyXL：Excel生成
    """)
